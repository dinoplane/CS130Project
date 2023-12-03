from openpyxl import Workbook, load_workbook

##first read contents of the file (replace with uploaded file)
uploaded_file = "00 37 33-excel.xlsx"
wb = load_workbook(uploaded_file)
ws = wb.active

#here"s where the modified data will be stored as a 2d list
modified_mapping = []
for i in range(1,ws.max_row+1):
    row = [cell.value for cell in ws[i]]
    isEmpty = True
    for col in row:
        if not ((col == None) or (col == "")):
            isEmpty = False
    if not isEmpty:
        modified_mapping.append(row)
print(wb.sheetnames[0])
print(modified_mapping)
