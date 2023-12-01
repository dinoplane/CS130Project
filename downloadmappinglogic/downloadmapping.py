from SPARQLWrapper import SPARQLWrapper, CSV, SPARQLExceptions
from openpyxl import Workbook
import csv
from io import StringIO


## WHAT YOU NEED TO GIVE



#### code starts here
def main():

    ##replace with user provided url
    fuseki_url = "http://localhost:3030/db/"


    sparql = SPARQLWrapper(fuseki_url)
    sparql.setReturnFormat(CSV)


    ##hardcoded query, replace with user provided query string
    user_provided_query =  """SELECT ?name ?loc
                WHERE {
                    ?subject start-date ?sd .
                    ?subject name ?name .
                    ?subject location ?loc .
                    FILTER(?sd < "2019-01-01")
                }
                LIMIT 25
    """    

    #flatten the query so we can save it in the file later 
    user_provided_query = " ".join(line.strip() for line in user_provided_query.splitlines())
    ## ^^^^ this is what backend needs to store in the downloaded excel file SAVE THIS QUERY
    ### do not forget!!!


    ## make a copy of that to modify for our logic
    query = user_provided_query

    #check to make sure it is NOT an update query for security reasons, and don't allow construct either
    for unallowed in ['INSERT', 'DELETE', 'LOAD', 'CLEAR', 'CONSTRUCT']:
        if unallowed in query:
            #if is an update query, we just download an empty excel file for now
            #update this behaviour however you like  <------------------------------------!!!!!
            wb = Workbook()
            ws = wb.active
            ## if we download, modify so that theres a download prompt for user.
            wb.save('not-allowed-query.xlsx')
            print("query not allowed")
            return


    #### here we take care of the prefix issue
    prefix = 'data:/'
    shorten = "a:"


    predicates = []
    get_predicates_query = "SELECT DISTINCT ?p WHERE {?s ?p ?o}ORDER BY ?p"

    #first we get our predicates
    sparql.setQuery(get_predicates_query)
    ret = sparql.queryAndConvert()
    content = ret.decode()
    predicates = [line.strip() for line in content.splitlines()][1:]
    predicates = [s.replace(prefix, '') for s in predicates]


    # flatten query for ease of parsing
    query = " ".join(line.strip() for line in query.splitlines())

    
    #add prefixes where necessary in query
    for pred in predicates:
        query = query.replace(pred, shorten+pred)
        query = query.replace('?' + shorten, '?')
    
    #and this is our query with all the prefixes added, pure SPARQL
    query = "PREFIX "+ shorten + " <" + prefix + "> " + query





    #We must now find if they SELECT the primary key too in their query
    #If not, we force them to. this is used later when we parse their
    #uploaded excel sheet for upload mapping(and sparql update)

    #first we find what the primary key is called in this query
    #its the first
    try:
        select_index = query.find("SELECT") + 6
        where_index = query.find("WHERE",select_index)
        curly_index = query.find("{", where_index)
        question_index = query.find("?", curly_index)

        #and here's out primary key's alias
        subject_name = query[question_index:].split()[0]

        #now we check if they user requests to display that alias
        #SELECT ?id ?name , ex: ?id is primary here
        chosen_labels = query[select_index:where_index].split()

        #if they don't select primary key, we force them to select it
        if subject_name not in chosen_labels:
            query = query[:select_index] + ' ' + subject_name + query[select_index:]

    except:
        #i dont wanna check all the edge cases for strings so i assume if 
        #we hit an error here the query was bad to begin with
        #again, replace this with whatever behaviour you do for bad queries
        wb = Workbook()
        ws = wb.active
        wb.save('bad-formed-query.xlsx')
        print("bad query formed")
        return
        
    #now lets query the kb and get the excel file
    ret = None
    try:  
        sparql.setQuery(query)
        ret = sparql.queryAndConvert()
    except(SPARQLExceptions.QueryBadFormed):
        # bad query, we must do smth to handle
        # for now jst download empty excel file, you'll have to change this
        wb = Workbook()
        ws = wb.active
        wb.save('bad-formed-query.xlsx')
        print("bad query fored")
        return

    content = ret.decode()

    #content is a string storing output. parse and remove prefixes in our query output.
    content = content.replace(prefix, "")

    #convert to inputstream
    file = StringIO(content)

    #read inputstream into a csv format
    csv_data = csv.reader(file, delimiter=",")

    #read csv into excel workbook
    wb = Workbook()
    ws = wb.active
    for r, row in enumerate(csv_data, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = val

    #rn we download the file our our directory, instead the user should be downloading it
    # don't forget to also store the user_provided_query in the file somehow
    wb.save('mapping-query-output.xlsx')


## main
if __name__ == "__main__":
    # you want to put the whole program in a try catch block
    # because fuseki might not be set up yet, or might be and invalid link
    try:
        main()
    except:
        #replace this with behaviour below in backend
        # if the program crashes for any reason, show the user
        # an error saying "something went wrong" or
        # "check if fuseki server is still connected"
        print("bad error")


