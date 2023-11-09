from SPARQLWrapper import SPARQLWrapper, CSV
from openpyxl import Workbook, load_workbook
import csv
from io import StringIO
import requests

def is_float(s):
    if s.replace(".", "").isnumeric():
        return True
    else:
        return False
    

def isInteger(num):
   

    val = int(num)
 
    diff = num - val

    if (diff > 0):
        return False
         
    return True

wb = load_workbook('myexcel.xlsx')
ws = wb.active

ttl_filestream = "@prefix a: <data:/> . \n\n"

headers = []

'''

for i in range(1, ws.max_row+1):
    row = [cell.value for cell in ws[i]]
    if(i == 1):
        headers = row
    else:
        for j in range(len(row)):
            predicate = row[j]
            if (not str(predicate).isnumeric()):
                predicate = "\"" + str(predicate) + "\""
            subject = headers[j]
            ttl_filestream +=("a:dataEntry" + str(i-1)+ " " +  "a:" + subject + " " + str(predicate) + " .\n")
'''

for i in range(1, ws.max_row+1):
    row = [cell.value for cell in ws[i]]
    if(i == 1):
        headers = row
    else:
        primary = row[0]
        for j in range(1, len(row)):
            predicate = row[j]
            if (not is_float(str(predicate))):
                predicate = "\"" + str(predicate) + "\""
            else:
                predicate = int(predicate)
            subject = headers[j]
            ttl_filestream +=("a:" + primary+ " " +  "a:" + subject + " " + str(predicate) + " .\n")


   # print(row)

print(ttl_filestream)

data = ttl_filestream
headers = {'Content-Type': 'text/turtle;charset=utf-8'}
r = requests.post('http://localhost:3030/db/data', data=data, headers=headers)


#todo: what about duplicate primary keys in og excel sheet?
#rodo: number input formatting, excel stores them as floats regardless