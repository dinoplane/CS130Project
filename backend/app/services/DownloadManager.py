import logging

from SPARQLWrapper import SPARQLWrapper, CSV, SPARQLExceptions
from fastapi import HTTPException
from openpyxl import Workbook
import csv
from io import StringIO

from starlette.responses import StreamingResponse

from app.database.fuseki_connection import Fuseki
from app.schema.mapping_schema import DownloadRequestSchema, MappingEntry

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DownloadManager:

    def __init__(self, request: DownloadRequestSchema):

        self.request: DownloadRequestSchema = request
        self.sparql_client = Fuseki.get_connection(request.fuseki_url)

    def download(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        for mapping in self.request.selected_mappings:
            user_provided_query = mapping.query
            # flattened the query
            user_provided_query = " ".join(line.strip() for line in user_provided_query.splitlines())
            query = user_provided_query
            # check if the mapping is a valid mapping
            self.check_valid_mapping(mapping, query)

            # Handling prefixes
            prefix = 'data:/'
            shorten = "a:"

            get_predicates_query = "SELECT DISTINCT ?p WHERE {?s ?p ?o}ORDER BY ?p"

            # Get the predicates
            self.sparql_client.setQuery(get_predicates_query)
            converted_query = self.sparql_client.queryAndConvert()
            content = converted_query.decode()
            predicates = [line.strip() for line in content.splitlines()][1:]
            predicates = [s.replace(prefix, '') for s in predicates]

            # flatten query for ease of parsing
            query = " ".join(line.strip() for line in query.splitlines())

            # add prefixes where necessary in query
            for predicate in predicates:
                query = query.replace(predicate, shorten + predicate)
                query = query.replace('?' + shorten, '?')

            # and this is our query with all the prefixes added, pure SPARQL
            query = "PREFIX " + shorten + " <" + prefix + "> " + query

            # Check if the primary key is selected if not then we force user to select
            try:
                select_index = query.find("SELECT") + 6
                where_index = query.find("WHERE", select_index)
                curly_index = query.find("{", where_index)
                question_index = query.find("?", curly_index)

                # Primary key's alias
                subject_name = query[question_index:].split()[0]

                # Check if they user requests to display that alias
                # SELECT ?id ?name , ex: ?id is primary here
                chosen_labels = query[select_index:where_index].split()

                # if they don't select primary key, we force them to select it
                if subject_name not in chosen_labels:
                    query = query[:select_index] + ' ' + subject_name + query[select_index:]

            except Exception as error:
                logger.error(str(error))
                raise HTTPException(status_code=500, detail=f"Primary key is not selected in the {mapping.query}")
            
            try:
                self.sparql_client.setQuery(query)
                ret = self.sparql_client.queryAndConvert()
            except SPARQLExceptions.QueryBadFormed as error:
                logger.error(str(error))
                raise HTTPException(status_code=500, detail=f"Bad query")

            content = ret.decode()

            # content is a string storing output. parse and remove prefixes in our query output.
            content = content.replace(prefix, "")

            # convert to inputstream
            file = StringIO(content)

            # read inputstream into a csv format
            csv_data = csv.reader(file, delimiter=",")

            # read csv into excel workbook
            sheet_name = f"sheet-{mapping.id}"
            sheet = workbook.create_sheet(title=sheet_name)

            for r, row in enumerate(csv_data, start=1):
                for c, val in enumerate(row, start=1):
                    sheet.cell(row=r, column=c).value = val

        content = save_workbook_to_memory(workbook)
        filename = "output.xlsx"
        headers = {"Content-Disposition": f"attachment; filename={filename}"}

        return StreamingResponse(content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                 , headers=headers)

    def check_valid_mapping(self, mapping: MappingEntry, query: str):
        for not_permitted_keyword in ['INSERT', 'DELETE', 'LOAD', 'CLEAR', 'CONSTRUCT']:
            if not_permitted_keyword in query:
                raise HTTPException(status_code=500, detail=f"The mapping {mapping.query} not permitted actions: "
                                                            f"'INSERT', 'DELETE', 'LOAD', 'CLEAR', 'CONSTRUCT'")


def save_workbook_to_memory(workbook):
    from io import BytesIO
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
