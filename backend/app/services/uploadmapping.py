from SPARQLWrapper import SPARQLWrapper, CSV, SPARQLExceptions
from openpyxl import Workbook, load_workbook
import csv
from io import StringIO


## WHAT YOU NEED TO GIVE

#### code starts here
def upload_mapping(fuseki_url, query, sheet):
    ##replace with user provided url

    sparql = SPARQLWrapper(fuseki_url)
    sparql.setReturnFormat(CSV)

    ###replace this with the query gotten from the excel sheet
    user_provided_query = query

    #flatten the query for parsing purposes. it should already be flattened because
    #downloadmapping flattens it for backend to store in excel file
    user_provided_query = " ".join(line.strip() for line in user_provided_query.splitlines())
    ##DO NOT MODIFY THIS VARIABLE ^^^ ######

    query = user_provided_query

    #### resolve prefix issue
    prefix = 'data:/'
    shorten = "a:"


    predicates = []
    get_predicates_query = "SELECT DISTINCT ?p WHERE {?s ?p ?o}ORDER BY ?p"

    #first we get our predicates
    sparql.setQuery(get_predicates_query)
    ret = sparql.queryAndConvert()
    content = ret.decode()
    predicates = [line.strip() for line in content.splitlines()][1:]
    predicates = [s.replace(prefix, '') for s in predicates]


    # flatten query for ease of parsing
    query = " ".join(line.strip() for line in query.splitlines())

    #add prefixes where necessary in query
    for pred in predicates:
        query = query.replace(pred, shorten+pred)
        query = query.replace('?' + shorten, '?')
    #and this is our query with all the prefixes added, pure SPARQL
    query = "PREFIX "+ shorten + " <" + prefix + "> " + query

    #We must now find if they SELECT the primary key too in their query
    #If not, we force them to. this is used later when we parse their
    #uploaded excel sheet for upload mapping(and sparql update)

    #first we find what the primary key is called in this query
    #its the first
    subject_name = None
    try:
        select_index = query.find("SELECT") + 6
        where_index = query.find("WHERE",select_index)
        curly_index = query.find("{", where_index)
        question_index = query.find("?", curly_index)

        #and here's out primary key's alias
        subject_name = query[question_index:].split()[0]

        #now we check if they user requests to display that alias
        #SELECT ?id ?name , ex: ?id is primary here
        chosen_labels = query[select_index:where_index].split()

        #if they don't select primary key, we force them to select it
        if subject_name not in chosen_labels:
            query = query[:select_index] + ' ' + subject_name + query[select_index:]

    except:
        wb = Workbook()
        worksheet = wb.active
        wb.save('bad-formed-query.xlsx')
        print("bad query formed")
        return


    #query the kb and get the excel file
    ret = None
    try:
        sparql.setQuery(query)
        ret = sparql.queryAndConvert()
    except(SPARQLExceptions.QueryBadFormed):
        # bad query, we must do smth to handle
        # for now jst download empty excel file, you'll have to change this
        wb = Workbook()
        worksheet = wb.active
        wb.save('bad-formed-query.xlsx')
        print("bad query fored")
        return

    content = ret.decode()

    #content is a string storing output. parse and remove prefixes in our query output.
    content = content.replace(prefix, "")

    #convert to inputstream
    file = StringIO(content)

    #read inputstream into a csv format
    original_csv_data = csv.reader(file, delimiter=",")


    ### original_mapping stores the original query in a 2d list
    original_mapping = []
    for row in original_csv_data:
        original_mapping.append(row)

    ############ now we read the modifed excel file into a 2d list
    worksheet = sheet

    # #here's where the modified data will be stored as a 2d list
    modified_mapping = []
    for i in range(1,worksheet.max_row+1):
        row = [cell.value for cell in worksheet[i]]
        isEmpty = True
        for col in row:
            if not ((col == None) or (col == '')):
                isEmpty = False
        if not isEmpty:
            modified_mapping.append(row)


    #first thing we do is check that the primary key exists in uploaded mapping
    pk = subject_name[1:]
    try:
        modified_mapping[0].index(pk)
    except:
        #there is no primary key in the uploaded mapping
        #therefore we can't know how to modify the kb
        print("no primary key in upload")
        return
    pk_index_mod = modified_mapping[0].index(pk)
    pk_index_orig = original_mapping[0].index(pk)

    #we also make sure that the two mappings have the same labels
    orig_labels = [i for i in original_mapping[0]]
    mod_labels = [i for i in modified_mapping[0]]
    orig_labels.sort()
    mod_labels.sort()
    if not (orig_labels == mod_labels):
        print("label mismatch")
        return

    #get relevant predicates

    tokenized = user_provided_query.split()
    #since labels can be an alias, we map the alias to the true label
    pred_map = {}
    try:
        for pred in predicates:
            if pred in tokenized:
                index = tokenized.index(pred)
                pred_map[tokenized[index+1][1:]] = pred
    except:
        print("bad query formed")
        return

    ## the first step is to see which rows to delete
    to_delete = []
    orig_keys = [i[pk_index_orig] for i in original_mapping]
    mod_keys = [i[pk_index_mod] for i in modified_mapping]

    for key in orig_keys:
        if key not in mod_keys:
            to_delete.append(key)

    #now the delete query to the fuseki kb
    values = "".join("a:" + id + " " for id in to_delete)
    delete_query = 'PREFIX a: <data:/> DELETE { ?s ?p ?o . } WHERE { VALUES ?s { ' + values + ' } ?s ?p ?o }'

    sparql.setQuery(delete_query)
    ret = sparql.queryAndConvert()


    #### now we have to handle inserts/modifications
    ## to modify a value, we must delete the triple and instert a new one
    ## to simply insert, we just insert.
    #below, we are deleting any common primary keys' relevant values from kb
    #and then inserting all of the primary keys' relevant values in modified mapping to kb
    triples = []
    for key in mod_keys:
        if (key == pk):
            continue
        key_index  = mod_keys.index(key)
        for label in pred_map:
            if label not in modified_mapping[0]:
                continue
            label_index = modified_mapping[0].index(label)
            val = modified_mapping[key_index][label_index]
            if not val or val == '':
                val = None
            elif (not str(val).replace(".", "").isnumeric()):
                val = "\"" + str(val) + "\""
            else:
                val = int(val)
            label = pred_map[label]
            triples.append([key, label, val])

    #we perform a delete-insert
    #first the delete
    remove_data = ""
    count = 0
    for entry in triples:
        if entry[0] not in orig_keys:
            continue
        remove_data +=("    a:" + str(entry[0])+ " " +  "a:" + str(entry[1]) + " ?a" + str(count) + " .\n")
        count+=1
    delete_query = "PREFIX a: <data:/> \nDELETE WHERE {\n" + remove_data + "}"

    sparql.setQuery(delete_query)
    ret = sparql.queryAndConvert()

    #now the insert
    add_data = ""
    for entry in triples:
        if not entry[2]:
            continue
        add_data +=("    a:" + str(entry[0])+ " " +  "a:" + str(entry[1]) + " " +str(entry[2]) + " .\n")
        count+=1
    insert_query = "PREFIX a: <data:/> \nINSERT DATA {\n" + add_data + "}"

    sparql.setQuery(insert_query)
    ret = sparql.queryAndConvert()